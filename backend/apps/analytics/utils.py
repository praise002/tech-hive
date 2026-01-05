import csv
from datetime import datetime
from io import BytesIO, StringIO

from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


class AnalyticsExporter:
    """Utility for exporting analytics data to CSV/Excel"""

    @staticmethod
    def export_dashboard_to_csv(data):
        """
        Export dashboard metrics to CSV format

        Args:
            data: Dashboard metrics dict from analytics_service

        Returns:
            HttpResponse with CSV file
        """
        output = StringIO()
        writer = csv.writer(output)

        writer.writerow(["TechHive Analytics Dashboard Export"])
        writer.writerow(["Period:", data["period"]])
        writer.writerow(
            [
                "Date Range:",
                f"{data['date_range']['start']} to {data['date_range']['end']}",
            ]
        )
        writer.writerow([])  # Blank row

        # Key Metrics Section
        writer.writerow(["KEY METRICS"])
        writer.writerow(["Metric", "Value", "Unit", "Change %", "Trend"])

        for metric_name, metric_data in data["metrics"].items():
            writer.writerow(
                [
                    metric_name.replace("_", " ").title(),
                    metric_data["value"],
                    metric_data["unit"],
                    metric_data["change_percentage"],
                    metric_data["trend"],
                ]
            )

        writer.writerow([])  # Blank row

        # Device Distribution Section
        writer.writerow(["DEVICE DISTRIBUTION"])
        writer.writerow(["Device", "Count", "Percentage"])

        for device in data["device_types"]:
            writer.writerow(
                [device["name"], device["value"], f"{device['percentage']}%"]
            )

        writer.writerow([])  # Blank row

        # Active Users Timeline Section
        writer.writerow(["ACTIVE USERS TIMELINE"])
        writer.writerow(
            ["Date", "Day", "Registered Users", "Visitors", "Total Active Users"]
        )

        for day_data in data["active_users"]:
            writer.writerow(
                [
                    day_data["date"],
                    day_data["day"],
                    day_data["registered_users"],
                    day_data["visitors"],
                    day_data["total_active_users"],
                ]
            )

        writer.writerow([])  # Blank row

        # Top Performing Posts Section
        writer.writerow(["TOP PERFORMING POSTS"])
        writer.writerow(["Category", "Title", "Views", "Shares", "ID"])

        for post in data["top_performing_posts"]:
            writer.writerow(
                [
                    post["category"],
                    post["title"],
                    post["views"],
                    post["shares"],
                    post["id"],
                ]
            )

        # Prepare response
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        filename = f"techhive_dashboard_{data['period']}_{datetime.now().strftime('%Y%m%d')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @staticmethod
    def export_dashboard_to_excel(data):
        """
        Export dashboard metrics to Excel format with styling

        Args:
            data: Dashboard metrics dict from analytics_service

        Returns:
            HttpResponse with Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "Dashboard Metrics"

        # Define styles
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)
        section_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )
        section_font = Font(bold=True, size=11)

        row = 1

        # Title
        ws[f"A{row}"] = "TechHive Analytics Dashboard Export"
        ws[f"A{row}"].font = Font(bold=True, size=14)
        row += 1

        # Metadata
        ws[f"A{row}"] = "Period:"
        ws[f"B{row}"] = data["period"]
        row += 1

        ws[f"A{row}"] = "Date Range:"
        ws[f"B{row}"] = f"{data['date_range']['start']} to {data['date_range']['end']}"
        row += 2

        # Key Metrics Section
        ws[f"A{row}"] = "KEY METRICS"
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Metric", "Value", "Unit", "Change %", "Trend"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for metric_name, metric_data in data["metrics"].items():
            ws[f"A{row}"] = metric_name.replace("_", " ").title()
            ws[f"B{row}"] = metric_data["value"]
            ws[f"C{row}"] = metric_data["unit"]
            ws[f"D{row}"] = metric_data["change_percentage"]
            ws[f"E{row}"] = metric_data["trend"]
            row += 1

        row += 1

        # Device Distribution Section
        ws[f"A{row}"] = "DEVICE DISTRIBUTION"
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Device", "Count", "Percentage"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for device in data["device_types"]:
            ws[f"A{row}"] = device["name"]
            ws[f"B{row}"] = device["value"]
            ws[f"C{row}"] = f"{device['percentage']}%"
            row += 1

        row += 1

        # Active Users Timeline Section
        ws[f"A{row}"] = "ACTIVE USERS TIMELINE"
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Date", "Day", "Registered Users", "Visitors", "Total Active Users"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for day_data in data["active_users"]:
            ws[f"A{row}"] = day_data["date"]
            ws[f"B{row}"] = day_data["day"]
            ws[f"C{row}"] = day_data["registered_users"]
            ws[f"D{row}"] = day_data["visitors"]
            ws[f"E{row}"] = day_data["total_active_users"]
            row += 1

        row += 1

        # Top Performing Posts Section
        ws[f"A{row}"] = "TOP PERFORMING POSTS"
        ws[f"A{row}"].fill = section_fill
        ws[f"A{row}"].font = section_font
        row += 1

        headers = ["Category", "Title", "Views", "Shares", "ID"]
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row += 1

        for post in data["top_performing_posts"]:
            ws[f"A{row}"] = post["category"]
            ws[f"B{row}"] = post["title"]
            ws[f"C{row}"] = post["views"]
            ws[f"D{row}"] = post["shares"]
            ws[f"E{row}"] = post["id"]
            row += 1

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Prepare response
        response = HttpResponse(
            output.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        filename = f"techhive_dashboard_{data['period']}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response

    @staticmethod
    def export_article_analytics_to_csv(data):
        """Export article analytics to CSV"""
        output = StringIO()
        writer = csv.writer(output)

        # Metadata
        writer.writerow(["TechHive Article Analytics Export"])
        writer.writerow(["Article ID:", data["article_id"]])
        writer.writerow(["Title:", data["title"]])
        writer.writerow(["Period:", data["period"]])
        writer.writerow(
            [
                "Date Range:",
                f"{data['date_range']['start']} to {data['date_range']['end']}",
            ]
        )
        writer.writerow([])

        # Metrics
        writer.writerow(["PERFORMANCE METRICS"])
        writer.writerow(["Metric", "Value"])
        writer.writerow(["Total Views", data["total_views"]])
        writer.writerow(["Unique Visitors", data["unique_visitors"]])
        writer.writerow(["Total Shares", data["total_shares"]])
        writer.writerow(["Avg Time on Page (min)", data["avg_time_on_page"]])
        writer.writerow(["Bounce Rate (%)", data["bounce_rate"]])

        # Prepare response
        response = HttpResponse(output.getvalue(), content_type="text/csv")
        filename = f"article_{data['article_id']}_{data['period']}_{datetime.now().strftime('%Y%m%d')}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'

        return response


analytics_exporter = AnalyticsExporter()
